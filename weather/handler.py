import asyncio
import os
import string
from typing import TYPE_CHECKING

import aioconsole as aio
import aiohttp
import openpyxl
from openpyxl.styles import Alignment, Font
from sqlalchemy import select

from config import get_settings
from core.database import async_session
from weather.models import WeatherData
from weather.schema import FactWeatherInfo, WeatherInfo

settings = get_settings()


if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession


class WeatherDataHandler:
    def __init__(self, api_key: str, meteo_url: str, db_period: int):
        self.api_key = api_key
        self.meteo_url = meteo_url
        self.db_period = db_period
        self.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.header_font = Font(bold=True)
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.headers = [
            "ID",
            "Date and time",
            "Temperature",
            "Wind Speed",
            "Wind Direction",
            "Pressure",
            "Precipitation Type",
            "Precipitation Amount",
        ]

    async def fetch_weather_data(self) -> FactWeatherInfo:
        """Возвращает данные о погоде в районе Сколтех."""
        headers = {"X-Yandex-Weather-Key": self.api_key}
        async with aiohttp.ClientSession() as session:
            async with session.get(url=self.meteo_url, headers=headers) as response:
                data = await response.json()
                return FactWeatherInfo(**data)

    async def save_weather_data(self, data: WeatherInfo) -> None:
        """Функция для сохранения данных в БД."""
        session: AsyncSession
        async with async_session() as session:
            record = WeatherData(
                temperature=data.temp,
                wind_speed=data.wind_speed,
                wind_direction=data.wind_dir,
                pressure=data.pressure_mm,
                precipitation_type=data.prec_type,
                precipitation_amount=data.prec_strength,
            )
            session.add(record)
            await session.commit()

    async def periodic_save_weather_data(self) -> None:
        """Периодическое сохранение записей о погоде в БД."""
        while True:
            weather_data = await self.fetch_weather_data()
            await self.save_weather_data(weather_data.fact)
            await asyncio.sleep(self.db_period)

    async def get_weather_data_limit10(self) -> list[WeatherData]:
        """Возвращает 10 последних записей таблицы WeatherData."""
        limit10_query = select(WeatherData).order_by(WeatherData.id.desc()).limit(10)
        session: AsyncSession
        async with async_session() as session:
            result = await session.execute(limit10_query)
            records: list[WeatherData] = result.scalars().all()
        return records

    def export_to_excel(self, weather_data: list[WeatherData]) -> None:
        """Функция для экспорта данных в Excel."""
        self.sheet.title = "Weather Data"
        # Заголовки столбцов
        self.sheet.append(self.headers)

        for cell in self.sheet[1]:
            cell.font = self.header_font

        for symb, header in zip(string.ascii_uppercase[: len(self.headers)], self.headers):
            self.sheet.column_dimensions[symb].width = (
                len(header) + 5 if header != "Date and time" else 20
            )

        for record in weather_data:
            self.sheet.append(
                [
                    record.id,
                    record.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    record.temperature,
                    record.wind_speed,
                    record.wind_direction,
                    record.pressure,
                    record.precipitation_type,
                    record.precipitation_amount,
                ]
            )

        # Выравнивание текста по центру для всех ячеек в колонках
        for row in self.sheet.iter_rows(
            min_row=1,
            max_row=self.sheet.max_row,
            min_col=1,
            max_col=self.sheet.max_column,
        ):
            for cell in row:
                cell.alignment = self.alignment

        output_file = os.path.join('doc', "weather_data.xlsx")
        self.workbook.save(output_file)
        print(f"Данные экспортированы в {output_file}")

    async def export_to_excel_with_input_command(self) -> None:
        """Экспорт данных в excel файл по вводу команды пользователем."""
        while True:
            command: str = await aio.ainput("Введите 'export' для экспорта данных в Excel: ")
            if command.strip().lower() == "export":
                weather_data = await self.get_weather_data_limit10()
                await asyncio.to_thread(self.export_to_excel, weather_data)


def get_weather_handler() -> WeatherDataHandler:
    return WeatherDataHandler(
        settings.API_WEATHER_KEY,
        settings.METEO_URL,
        settings.DB_PERIOD,
    )
